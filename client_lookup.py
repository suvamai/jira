"""
Client lookup from Excel file (NEW LR CLIENT LIST 2026.xlsx).

Enriches billable_account, db_code, list_manager from the client database.

Lookup order:
  1. Broker-specific sheet (e.g. AMLC sheet) — match rental_name against
     list_name then mailer_name (higher precision)
  2. Full client sheet — fuzzy name match on rental_name / db_name (fallback)
"""

import re
import logging
from pathlib import Path

log = logging.getLogger(__name__)

_BROKER_EXCEL   = Path(__file__).parent / "NEW LR CLIENT LIST 2026.xlsx"
_FALLBACK_EXCEL = Path(__file__).parent / "broker_pdf" / "NEW LR CLIENT LIST 2026 1.xlsx"

# Map list_manager values → sheet name in the broker Excel
_MANAGER_TO_SHEET = {
    "AMLC":             "AMLC",
    "ADSTRA":           "ADSTRA",
    "AALC":             "AALC",
    "CELCO":            "CELCO",
    "CONRAD":           "CONRAD",
    "DATA-AXLE":        "DATA-AXLE",
    "KAP":              "KAP",
    "MARY E GRANGER":   "MARY E GRANGER",
    "NEGEV":            "NEGEV",
    "NAMES IN THE NEWS":"NITN",
    "RKD":              "RKD",
    "RMI":              "RMI",
    "WASHINGTON LISTS": "WASHINGTON LIST",
    "WE ARE MOORE":     "WE ARE MOORE",
}

_sheet_cache:   dict[str, list[dict]] = {}
_client_cache:  list | None = None
_WORD_CLEAN_RE = re.compile(r"[^a-z0-9 ]")


def _words(s: str) -> set:
    return {w for w in _WORD_CLEAN_RE.sub(" ", s.lower()).split() if len(w) > 2}


def _word_overlap(a: str, b: str) -> float:
    wa, wb = _words(a), _words(b)
    if not wa:
        return 0.0
    matches = sum(
        1 for w in wa
        if any(w == v or w.startswith(v) or v.startswith(w) for v in wb)
    )
    return matches / len(wa)


def _clean_billing(raw: str) -> str:
    """Strip parenthetical suffixes: 'T11 (A42D)' → 'T11'."""
    return re.sub(r"\s*\(.*?\)", "", raw).strip()


def _load_broker_sheet(list_manager: str) -> list[dict]:
    """Load the broker-specific sheet from NEW LR CLIENT LIST 2026.xlsx. Cached."""
    sheet_name = _MANAGER_TO_SHEET.get((list_manager or "").upper().strip())
    if not sheet_name:
        return []
    if sheet_name in _sheet_cache:
        return _sheet_cache[sheet_name]

    if not _BROKER_EXCEL.exists():
        log.warning("Broker Excel not found: %s", _BROKER_EXCEL)
        _sheet_cache[sheet_name] = []
        return []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(_BROKER_EXCEL), read_only=True)
        if sheet_name not in wb.sheetnames:
            log.warning("Sheet %r not found in %s", sheet_name, _BROKER_EXCEL.name)
            _sheet_cache[sheet_name] = []
            return []

        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, max_col=8, values_only=True):
            db_code     = str(row[0] or "").strip()
            billing_raw = str(row[1] or "").strip()
            db_name     = str(row[2] or "").strip()
            rental_name = str(row[3] or "").strip()
            lm          = str(row[4] or "").strip()
            lm_contact  = str(row[5] or "").strip()
            if db_code:
                rows.append({
                    "db_code":      db_code,
                    "billing_cust": _clean_billing(billing_raw),
                    "db_name":      db_name,
                    "rental_name":  rental_name,
                    "list_manager": lm,
                    "lm_contact":   lm_contact,
                })
        wb.close()
        log.info("Loaded %d rows from sheet %r", len(rows), sheet_name)
        _sheet_cache[sheet_name] = rows
        return rows

    except Exception as e:
        log.warning("Failed to load broker sheet %r: %s", sheet_name, e)
        _sheet_cache[sheet_name] = []
        return []


def _load_all_clients(excel_path: str = None) -> list[dict]:
    """Load full client list (LIST RENTAL FULL CLIENT SHEET). Cached."""
    global _client_cache
    if _client_cache is not None:
        return _client_cache

    path = excel_path or str(_BROKER_EXCEL if _BROKER_EXCEL.exists() else _FALLBACK_EXCEL)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        clients = []
        for row in ws.iter_rows(min_row=2, max_col=8, values_only=True):
            db_code     = str(row[0] or "").strip()
            billing_raw = str(row[1] or "").strip()
            db_name     = str(row[2] or "").strip()
            rental_name = str(row[3] or "").strip()
            lm          = str(row[4] or "").strip()
            lm_contact  = str(row[5] or "").strip()
            if db_code or billing_raw:
                clients.append({
                    "db_code":      db_code,
                    "billing_cust": _clean_billing(billing_raw),
                    "db_name":      db_name,
                    "rental_name":  rental_name,
                    "list_manager": lm,
                    "lm_contact":   lm_contact,
                })
        wb.close()
        log.info("Loaded %d clients from full sheet", len(clients))
        _client_cache = clients
        return _client_cache
    except Exception as e:
        log.warning("Failed to load client Excel: %s", e)
        _client_cache = []
        return []


def _best_match(rows: list[dict], *names: str) -> tuple[dict | None, float]:
    """Return (best_row, best_score) by matching names against rental_name and db_name."""
    best, best_score = None, 0.0
    for row in rows:
        for name in names:
            if not name:
                continue
            for field in ("rental_name", "db_name"):
                score = _word_overlap(name, row.get(field, ""))
                if score > best_score:
                    best_score, best = score, row
    return best, best_score


def _row_to_result(row: dict) -> dict:
    return {
        "billable_account": row["billing_cust"],
        "list_manager":     row["list_manager"],
        "db_code":          row["db_code"],
        "lm_contact":       row["lm_contact"],
    }


def enrich_fields(
    list_name:    str = "",
    mailer_name:  str = "",
    list_manager: str = "",
    db_code:      str = "",
) -> dict:
    """
    Look up db_code, billable_account, and list_manager from Excel.

    Priority:
      1. Exact db_code match in broker sheet or full sheet
      2. Broker-specific sheet: fuzzy match on list_name, then mailer_name (threshold 0.4)
      3. Full client sheet: fuzzy match on list_name (threshold 0.5)

    Returns dict with billable_account, list_manager, db_code, lm_contact.
    Empty dict if no match found.
    """
    # 1. Exact db_code match
    if db_code:
        for row in _load_broker_sheet(list_manager) + _load_all_clients():
            if row["db_code"].upper().strip() == db_code.upper().strip():
                return _row_to_result(row)

    # 2. Broker-specific sheet — match list_name then mailer_name
    broker_rows = _load_broker_sheet(list_manager)
    if broker_rows:
        best, score = _best_match(broker_rows, list_name, mailer_name)
        if best and score >= 0.4:
            log.info("Broker sheet match (score=%.2f): %s → %s", score, list_name or mailer_name, best["db_code"])
            return _row_to_result(best)

    # 3. All other broker sheets (cross-broker fallback)
    for mgr_key, sheet_name in _MANAGER_TO_SHEET.items():
        if mgr_key == (list_manager or "").upper().strip():
            continue  # already tried
        rows = _load_broker_sheet(mgr_key)
        if not rows:
            continue
        best, score = _best_match(rows, list_name, mailer_name)
        if best and score >= 0.5:
            log.info("Cross-broker sheet %r match (score=%.2f): %s → %s", sheet_name, score, list_name or mailer_name, best["db_code"])
            return _row_to_result(best)

    # 4. Full client sheet fallback
    best, score = _best_match(_load_all_clients(), list_name)
    if best and score >= 0.5:
        log.info("Full sheet match (score=%.2f): %s → %s", score, list_name, best["db_code"])
        return _row_to_result(best)

    return {}


# Keep old signature working
def get_billable_account(list_name: str = "", db_code: str = "") -> dict:
    return enrich_fields(list_name=list_name, db_code=db_code)
